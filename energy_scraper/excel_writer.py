"""
excel_writer.py — Enterprise Excel report with 7 styled sheets + AI feedback system.

Features:
  • Color-coded rows by confidence band (Green / Yellow / Orange)
  • "AI Correct?" dropdown column on Output and Rejected sheets
  • "User Feedback for AI Learning" sheet for free-form feedback
  • Auto-filter on all data sheets
  • Styled headers (blue with white bold text)
  • Auto-width columns with freeze panes
  • Hyperlinked article URLs
  • Mega-deal highlighting (>$1B)
"""

import re
from pathlib import Path
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from .project_paths import RUN_REPORTS_DIR, ensure_runtime_dirs
except ImportError:
    from project_paths import RUN_REPORTS_DIR, ensure_runtime_dirs


# ── Colour palette ──
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MEGA_DEAL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Blue for >$1B
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FEEDBACK_HEADER = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
# Pale yellow fill for user-editable feedback & comments cells
FEEDBACK_INPUT = PatternFill(start_color="FFFFC0", end_color="FFFFC0", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
BODY_FONT   = Font(size=10, name="Calibri")
LINK_FONT   = Font(color="0563C1", underline="single", size=10, name="Calibri")
MEGA_FONT   = Font(bold=True, size=10, name="Calibri", color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _is_mega_deal(value_str: str) -> bool:
    """Check if deal value exceeds $1 Billion."""
    if not value_str or value_str == "Undisclosed":
        return False
    v = value_str.lower()
    if "billion" in v or "bn" in v:
        m = re.search(r'[\d,.]+', v)
        if m:
            try:
                num = float(m.group().replace(",", ""))
                return num >= 1.0
            except ValueError:
                pass
    if "trillion" in v or "tn" in v:
        return True
    return False


class ExcelReportWriter:

    OUTPUT_COLS = [
        "Headline", "Buyer", "Seller", "Asset", "Date",
        "Industry", "Sector", "Link", "Source", "Geography", "Value",
        "Processing Date", "Deal Type", "Deal Status", "Strategic Rationale",
        "CONFIDENCE THRESHOLDS", "AI Correct?",
        "Feedback to AI",  # User explains why decision was wrong — read by AI next run
        "Comments",        # General user notes (not fed to AI)
    ]

    LOG_COLS = [
        "#", "Website", "Section / Sub-section",
        "Fetch Method Used", "Access Mode", "Rendering Type",
        "Status", "Total Articles Found", "Articles in Date Range",
        "M&A Deals Extracted (Sheet 1)", "Review Queue Items (Sheet 2)",
        "Issues Encountered", "Resolution Applied",
    ]

    ISSUE_COLS = [
        "#", "Issue Category", "Description", "Affected Website(s)",
        "Severity", "Solution Applied This Run", "Recommended Permanent Fix",
    ]

    REJECTED_COLS = [
        "Headline", "Buyer", "Seller", "Asset", "Date",
        "Industry", "Sector", "Link", "Source", "Geography", "Value",
        "Deal Type", "Deal Status", "Strategic Rationale",
        "Confidence", "Rejection Reason", "AI Correct?",
        "Feedback to AI",  # User explains why AI was wrong — read by AI next run
        "Comments",        # General notes
    ]

    FEEDBACK_COLS = [
        "Feedback / Instruction for AI",
        "Type (Accept Rule / Reject Rule / Company Name / Other)",
        "Related Headline (optional)",
        "Date Added",
    ]

    def __init__(self):
        timestr = datetime.now().strftime("%Y%m%d_%H%M%S")
        ensure_runtime_dirs()
        self.default_filename = str(RUN_REPORTS_DIR / f"Energy_MA_Report_Async_{timestr}.xlsx")

    @staticmethod
    def _row_fill(confidence: float, value_str: str = ""):
        if _is_mega_deal(value_str):
            return MEGA_DEAL  # Special blue for mega-deals
        if confidence >= 0.80:
            return GREEN
        elif confidence >= 0.50:
            return YELLOW
        return ORANGE

    def export(self, deals, logs, issues, metrics, filename=None, rejected_deals=None, ai_stats=None, run_summary=None):
        output_path = Path(filename or self.default_filename)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        filename = output_path
        now = datetime.now().strftime("%Y-%m-%d")
        rejected_deals = rejected_deals or []

        # ── Split deals by industry sheet ──
        INDUSTRY_SHEETS = ["Upstream", "Midstream", "OFS", "R&M", "P&U", "JV & Partnerships"]
        REPORTS_COLS = ["Headline", "Link", "Date"]

        industry_deals = {s: [] for s in INDUSTRY_SHEETS}
        report_items = []
        review_deals = []

        for d in deals:
            sheet = d.get("Sheet", "P&U")
            conf = d.get("Confidence", 0.0)
            if sheet == "Reports":
                report_items.append(d)
            elif conf < 0.50:
                review_deals.append(d)
            else:
                industry_deals.setdefault(sheet, []).append(d)

        def _deal_rows(deal_list, include_ai_correct=True):
            rows = []
            for d in deal_list:
                row = {
                    "Headline":              d.get("Headline", ""),
                    "Buyer":                 d.get("Buyer", ""),
                    "Seller":                d.get("Seller", ""),
                    "Asset":                 d.get("Asset", ""),
                    "Date":                  d.get("Date", ""),
                    "Industry":              d.get("Industry", ""),
                    "Sector":                d.get("Sector", ""),
                    "Link":                  d.get("Link", ""),
                    "Source":                d.get("Source", ""),
                    "Geography":             d.get("Geography", d.get("County", "Global")),
                    "Value":                 d.get("Value", "Undisclosed"),
                    "Processing Date":       now,
                    "Deal Type":             d.get("Deal Type", ""),
                    "Deal Status":           d.get("Deal Status", "Announced"),
                    "Strategic Rationale":   d.get("Strategic Rationale", "N/A"),
                    "CONFIDENCE THRESHOLDS": d.get("Confidence", 0.0),
                }
                if include_ai_correct:
                    row["AI Correct?"] = ""    # User fills: Yes / No
                    row["Feedback to AI"] = "" # User explains if No
                    row["Comments"] = ""       # General notes
                rows.append(row)
            return rows

        # Build DataFrames for each industry sheet
        industry_dfs = {}
        for sheet_name in INDUSTRY_SHEETS:
            sheet_deals = industry_deals.get(sheet_name, [])
            industry_dfs[sheet_name] = (
                pd.DataFrame(_deal_rows(sheet_deals), columns=self.OUTPUT_COLS)
                if sheet_deals else pd.DataFrame(columns=self.OUTPUT_COLS)
            )

        df_rev  = pd.DataFrame(_deal_rows(review_deals), columns=self.OUTPUT_COLS) if review_deals else pd.DataFrame(columns=self.OUTPUT_COLS)
        df_log  = pd.DataFrame(logs,   columns=self.LOG_COLS)   if logs   else pd.DataFrame(columns=self.LOG_COLS)
        df_iss  = pd.DataFrame(issues, columns=self.ISSUE_COLS) if issues else pd.DataFrame(columns=self.ISSUE_COLS)

        # Reports sheet: only Headline, Link, Date
        report_rows = [{"Headline": d.get("Headline", ""), "Link": d.get("Link", ""), "Date": d.get("Date", "")} for d in report_items]
        df_reports = pd.DataFrame(report_rows, columns=REPORTS_COLS) if report_rows else pd.DataFrame(columns=REPORTS_COLS)

        # Add AI stats to metrics
        if ai_stats:
            metrics["AI Provider Stats"] = ai_stats
        df_dash = pd.DataFrame(list(metrics.items()), columns=["Metric", "Value"]) if metrics else pd.DataFrame(columns=["Metric", "Value"])

        # Rejected by AI sheet
        rej_rows = []
        for d in rejected_deals:
            rej_rows.append({
                "Headline":            d.get("Headline", ""),
                "Buyer":               d.get("Buyer", ""),
                "Seller":              d.get("Seller", ""),
                "Asset":               d.get("Asset", ""),
                "Date":                d.get("Date", ""),
                "Industry":            d.get("Industry", ""),
                "Sector":              d.get("Sector", ""),
                "Link":                d.get("Link", ""),
                "Source":              d.get("Source", ""),
                "Geography":           d.get("Geography", d.get("County", "Global")),
                "Value":               d.get("Value", "Undisclosed"),
                "Deal Type":           d.get("Deal Type", ""),
                "Deal Status":         d.get("Deal Status", "Rejected/Other"),
                "Strategic Rationale": d.get("Strategic Rationale", "N/A"),
                "Confidence":          d.get("Confidence", 0.0),
                "Rejection Reason":    d.get("Rejection Reason", ""),
                "AI Correct?":         "",
                "Feedback to AI":      "",
                "Comments":            "",
            })
        df_rej = pd.DataFrame(rej_rows, columns=self.REJECTED_COLS) if rej_rows else pd.DataFrame(columns=self.REJECTED_COLS)

        # User Feedback sheet (starts empty, user fills in)
        df_feedback = pd.DataFrame(columns=self.FEEDBACK_COLS)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            wb = writer.book
            # ── ARIA Run Intelligence (New) ──
            if run_summary:
                ws_summary = wb.create_sheet("Run Intelligence", 0)
                ws_summary.append(["ARIA EXECUTIVE RUN SUMMARY"])
                ws_summary.append(["Generated on: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                ws_summary.append([""])
                
                # Split summary into sentences for better display
                sentences = [s.strip() + "." for s in run_summary.split(".") if s.strip()]
                for s in sentences:
                    ws_summary.append([s])
                
                # Styling
                ws_summary["A1"].font = Font(bold=True, size=14, color="1F4E79")
                ws_summary["A2"].font = Font(italic=True, size=10, color="595959")
                
                for row in ws_summary.iter_rows(min_row=4, max_row=ws_summary.max_row):
                    for cell in row:
                        cell.font = Font(size=12, name="Calibri")
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                ws_summary.column_dimensions["A"].width = 120

            # ── Industry sheets ──
            for sheet_name in INDUSTRY_SHEETS:
                industry_dfs[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
            # ── Reports sheet ──
            df_reports.to_excel(writer, sheet_name="Reports", index=False)
            # ── Existing utility sheets ──
            df_rev.to_excel(writer,      sheet_name="Review Queue",                  index=False)
            df_rej.to_excel(writer,      sheet_name="Rejected by AI",                index=False)
            df_log.to_excel(writer,      sheet_name="Website Processing Log",        index=False)
            df_iss.to_excel(writer,      sheet_name="Issue and Solution",            index=False)
            df_dash.to_excel(writer,     sheet_name="Summary Dashboard",             index=False)
            df_feedback.to_excel(writer, sheet_name="User Feedback for AI Learning", index=False)

            wb = writer.book

            # ── "AI Correct?" dropdown validation ──
            ai_validation = DataValidation(
                type="list",
                formula1='"Yes,No"',
                allow_blank=True,
                showDropDown=False,
            )
            ai_validation.error = "Please select Yes or No"
            ai_validation.errorTitle = "AI Feedback"
            ai_validation.prompt = "Was the AI's decision correct? Select Yes or No"
            ai_validation.promptTitle = "AI Feedback"

            # ── Feedback type dropdown ──
            feedback_type_validation = DataValidation(
                type="list",
                formula1='"Accept Rule,Reject Rule,Company Name,Other"',
                allow_blank=True,
                showDropDown=False,
            )

            # ── Style industry deal sheets + Review Queue ──
            styled_sheets = INDUSTRY_SHEETS + ["Review Queue"]
            for sheet_name in styled_sheets:
                ws = wb[sheet_name]
                self._style_headers(ws)
                self._freeze_panes(ws)

                conf_col = self.OUTPUT_COLS.index("CONFIDENCE THRESHOLDS") + 1
                link_col = self.OUTPUT_COLS.index("Link") + 1
                ai_col = self.OUTPUT_COLS.index("AI Correct?") + 1
                value_col = self.OUTPUT_COLS.index("Value") + 1
                rationale_col = self.OUTPUT_COLS.index("Strategic Rationale") + 1
                feedback_col = self.OUTPUT_COLS.index("Feedback to AI") + 1
                comments_col = self.OUTPUT_COLS.index("Comments") + 1

                # Add AI Correct? dropdown to all data rows
                ai_val = DataValidation(
                    type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False,
                )
                ai_val.prompt = "Was the AI's decision correct? Select Yes or No"
                ws.add_data_validation(ai_val)
                ai_col_letter = get_column_letter(ai_col)
                if ws.max_row >= 2:
                    ai_val.add(f"{ai_col_letter}2:{ai_col_letter}{ws.max_row}")

                for row_idx in range(2, ws.max_row + 1):
                    conf_val = ws.cell(row=row_idx, column=conf_col).value
                    value_str = str(ws.cell(row=row_idx, column=value_col).value or "")
                    try:
                        conf_val = float(conf_val)
                    except (TypeError, ValueError):
                        conf_val = 0.0
                    fill = self._row_fill(conf_val, value_str)

                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        # Feedback to AI and Comments get pale yellow to indicate user-editable
                        if col_idx in (feedback_col, comments_col):
                            cell.fill = FEEDBACK_INPUT
                        else:
                            cell.fill = fill
                        cell.font = MEGA_FONT if _is_mega_deal(value_str) else BODY_FONT
                        cell.border = THIN_BORDER

                    # Hyperlink
                    link_cell = ws.cell(row=row_idx, column=link_col)
                    url = str(link_cell.value or "")
                    if url.startswith("http"):
                        link_cell.hyperlink = url
                        link_cell.font = LINK_FONT

                # Set wider widths for the rationale and user-input columns
                ws.column_dimensions[get_column_letter(rationale_col)].width = 50
                ws.column_dimensions[get_column_letter(feedback_col)].width = 55
                ws.column_dimensions[get_column_letter(comments_col)].width = 40
                self._auto_width(ws)
                # Auto-filter
                ws.auto_filter.ref = ws.dimensions

            # ── Style Reports sheet ──
            ws_rep = wb["Reports"]
            self._style_headers(ws_rep)
            self._freeze_panes(ws_rep)
            link_col_rep = REPORTS_COLS.index("Link") + 1
            for row_idx in range(2, ws_rep.max_row + 1):
                for col_idx in range(1, ws_rep.max_column + 1):
                    cell = ws_rep.cell(row=row_idx, column=col_idx)
                    cell.font = BODY_FONT
                    cell.border = THIN_BORDER
                link_cell = ws_rep.cell(row=row_idx, column=link_col_rep)
                url = str(link_cell.value or "")
                if url.startswith("http"):
                    link_cell.hyperlink = url
                    link_cell.font = LINK_FONT
            self._auto_width(ws_rep)
            ws_rep.auto_filter.ref = ws_rep.dimensions

            # ── Style Rejected sheet with AI dropdown ──
            ws_rej = wb["Rejected by AI"]
            self._style_headers(ws_rej)
            self._freeze_panes(ws_rej)
            
            ai_col_rej = self.REJECTED_COLS.index("AI Correct?") + 1
            feedback_col_rej = self.REJECTED_COLS.index("Feedback to AI") + 1
            comments_col_rej = self.REJECTED_COLS.index("Comments") + 1
            ai_rej_validation = DataValidation(
                type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False,
            )
            ai_rej_validation.prompt = "Was AI correct to reject this? No = AI should have accepted it"
            ws_rej.add_data_validation(ai_rej_validation)
            ai_rej_letter = get_column_letter(ai_col_rej)
            if ws_rej.max_row >= 2:
                ai_rej_validation.add(f"{ai_rej_letter}2:{ai_rej_letter}{ws_rej.max_row}")
            
            for row_idx in range(2, ws_rej.max_row + 1):
                for col_idx in range(1, ws_rej.max_column + 1):
                    cell = ws_rej.cell(row=row_idx, column=col_idx)
                    if col_idx in (feedback_col_rej, comments_col_rej):
                        cell.fill = FEEDBACK_INPUT
                    cell.font = BODY_FONT
                    cell.border = THIN_BORDER

            ws_rej.column_dimensions[get_column_letter(feedback_col_rej)].width = 55
            ws_rej.column_dimensions[get_column_letter(comments_col_rej)].width = 40
            self._auto_width(ws_rej)
            ws_rej.auto_filter.ref = ws_rej.dimensions

            # ── Style remaining sheets ──
            for sheet_name in ["Website Processing Log", "Issue and Solution", "Summary Dashboard"]:
                ws = wb[sheet_name]
                self._style_headers(ws)
                self._freeze_panes(ws)
                self._auto_width(ws)
                ws.auto_filter.ref = ws.dimensions
                for row_idx in range(2, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = BODY_FONT
                        cell.border = THIN_BORDER

            # ── Style Feedback sheet ── 
            ws_fb = wb["User Feedback for AI Learning"]
            self._style_headers(ws_fb, fill=FEEDBACK_HEADER)
            self._freeze_panes(ws_fb)
            
            # Add feedback type dropdown
            ws_fb.add_data_validation(feedback_type_validation)
            type_col_letter = get_column_letter(2)  # Column B = Type
            feedback_type_validation.add(f"{type_col_letter}2:{type_col_letter}100")
            
            # Add 20 empty rows for user to fill in
            for row_idx in range(2, 22):
                ws_fb.cell(row=row_idx, column=4).value = datetime.now().strftime("%Y-%m-%d")
                for col_idx in range(1, len(self.FEEDBACK_COLS) + 1):
                    cell = ws_fb.cell(row=row_idx, column=col_idx)
                    cell.font = BODY_FONT
                    cell.border = THIN_BORDER

            # Set wider columns for feedback sheet
            ws_fb.column_dimensions["A"].width = 60
            ws_fb.column_dimensions["B"].width = 35
            ws_fb.column_dimensions["C"].width = 50
            ws_fb.column_dimensions["D"].width = 15

        total_industry = sum(len(v) for v in industry_deals.values())
        print(f"\n  ✅ Report saved → {filename}")
        for sn in INDUSTRY_SHEETS:
            cnt = len(industry_deals.get(sn, []))
            if cnt:
                print(f"     {sn:12s} deals: {cnt}")
        print(f"     Total deals:        {total_industry}")
        print(f"     Reports:            {len(report_items)}")
        print(f"     Review Queue items: {len(review_deals)}")
        print(f"     Rejected by AI:     {len(rejected_deals)}")
        if ai_stats:
            print(f"     AI Stats:           {ai_stats}")

    # ── Styling helpers ──

    @staticmethod
    def _style_headers(ws, fill=None):
        fill = fill or HEADER_FILL
        for cell in ws[1]:
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    @staticmethod
    def _freeze_panes(ws):
        ws.freeze_panes = "A2"

    @staticmethod
    def _auto_width(ws, max_width=55):
        for col_cells in ws.columns:
            lengths = [len(str(c.value or "")) for c in col_cells]
            best = min(max(lengths) + 3, max_width)
            ws.column_dimensions[col_cells[0].column_letter].width = best
