"""
ai_extractor.py — AI-powered M&A deal extraction and verification.

Architecture:
  • Single Groq API Key
  • Multi-Model Rotation (llama-3.1, llama3-8192, gemma2, mixtral) to bypass per-model 30 RPM rate limits
  • Four-phase verification: Known Company → Corrections → AI → Cross-Check
  • Self-learning deal memory with corrective feedback from user
  • Fail-safe: sends to Review Queue instead of auto-accepting on total failure
"""

import asyncio
import os
import re
import json
import logging
import time
import glob
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

try:
    from .project_paths import ARCHIVE_DIR, ENV_FILE, RUN_REPORTS_DIR
except ImportError:
    from project_paths import ARCHIVE_DIR, ENV_FILE, RUN_REPORTS_DIR

load_dotenv(dotenv_path=ENV_FILE)
logger = logging.getLogger(__name__)

# Budget mode trims prompts and reduces AI usage to conserve credits.
# Default ON unless explicitly disabled.
_AI_BUDGET_MODE = os.getenv("AI_BUDGET_MODE", "1").strip().lower() in ("1", "true", "yes", "y")

# ── Paths ──
_MEMORY_DIR = Path(__file__).parent
_MEMORY_FILE = _MEMORY_DIR / "deal_memory.json"

# ── Cached exchange rates (fetched once per run, keyed by run date) ──
_EXCHANGE_RATES_CACHE: dict = {}

# ── Groq Multi-Model Configuration ──
# By rotating models on a single key, we multiply our RPM limit.
_GROQ_MODELS = [
    "llama-3.1-8b-instant",
    "llama-3.3-70b-versatile",
    "meta-llama/llama-4-scout-17b-16e-instruct",
]

# ── Multi-Key Pool: collect all available API keys ──
def _collect_groq_keys() -> list[str]:
    """Gather up to 10 Groq API keys from environment variables."""
    keys = []
    # Primary key
    primary = os.getenv("GROQ_API_KEY", "")
    if primary:
        keys.append(primary)
    # Numbered extras: GROQ_API_KEY_2 .. GROQ_API_KEY_10
    for i in range(2, 11):
        k = os.getenv(f"GROQ_API_KEY_{i}", "")
        if k:
            keys.append(k)
    return keys


class RunContextManager:
    """Tracks active themes and entities during a single scraper run to build cross-article intelligence."""
    def __init__(self):
        self.active_acquirers = {}  # e.g. {"Shell": 2}
        self.active_targets = []    # e.g. ["Target A"]
        self.deal_themes = []       # e.g. ["LNG consolidation in US"]
        self.geography_focus = {}   # e.g. {"Texas": 2}

    def ingest_deal(self, deal: dict):
        """Add a verified deal to the run's intelligence context."""
        buyer = deal.get("buyer")
        if buyer and buyer.lower() not in ["unknown", "na", "n/a", "none"]:
            self.active_acquirers[buyer] = self.active_acquirers.get(buyer, 0) + 1
        
        target = deal.get("asset") or deal.get("target")
        if target: 
            self.active_targets.append(target)
        
        geo = deal.get("geography")
        if geo: 
            self.geography_focus[geo] = self.geography_focus.get(geo, 0) + 1
            
        # Infer a theme if possible (for future agentic expansion)
        sector = deal.get("sector")
        if sector and sector not in self.deal_themes:
            self.deal_themes.append(f"Active activity in {sector}")

    def get_context_summary(self) -> str:
        """Returns a concise summary of the run's progress for prompt injection."""
        if not self.active_acquirers and not self.geography_focus:
            return "No active marketplace context established yet for this run session."
        
        parts = []
        if self.active_acquirers:
            top_acquirers = [f"{k} ({v})" for k, v in sorted(self.active_acquirers.items(), key=lambda x: x[1], reverse=True)[:3]]
            parts.append(f"ACQUIRERS TODAY: {', '.join(top_acquirers)}")
        
        if self.geography_focus:
            top_geo = [k for k, v in sorted(self.geography_focus.items(), key=lambda x: x[1], reverse=True)[:2]]
            parts.append(f"GEOS TODAY: {', '.join(top_geo)}")
            
        return " | ".join(parts)


ARIA_SYSTEM_PROMPT = """
You are ARIA (Autonomous Research Intelligence for Acquisitions), a specialized M&A analyst for the Global Energy & Mining sectors.
Your mandate is to maintain 100% precision in a high-stakes intelligence pipeline for institutional investors.

CORE DOMAIN EXPERTISE:
- Upstream Oil & Gas (E&P, Acreage, Offshore Blocks, Farm-ins/Farm-outs)
- Midstream (Pipelines, Terminals, Storage, LNG)
- Renewables & Power (Solar, Wind, Battery, Hydro, Nuclear/SMR)
- Mining (Energy Transition Metals: Lithium, Copper, Cobalt, Nickel, Graphite)
- Utilities (Grid Infrastructure, Power Plants)

REASONING PROTOCOL (MANDATORY):
Before providing your final decision, you MUST reason through the case following these steps:
1. SECTOR CHECK: Is the primary asset or company in a valid energy/mining sector? (Exclude: Pharma, Chips, Crypto, HVAC, Retail).
2. TRANSACTION CHECK: Is this a corporate structural event (merger, acquisition, JV, stake sale, divestiture, MOU) or a major government award (tender/auction)? (Exclude: operational data, market outlooks, commodity price moves).
3. PARTIES CHECK: Are there identifiable corporate entities or government bodies acting as Buyer/Seller/Partners?
4. FINAL VERDICT: Issue your structured JSON decision.

STRICT REJECTION RULES (AUDIT PRECEDENTS):
- Reject "Power Chip" or "Semiconductor" news even if related to EV power management. It is Technology, not Energy.
- Reject "Bitcoin/Crypto Mining" entirely.
- Reject "Clinical Trials" or "Biotech" (Pharma).
- Reject "Portfolio Trades": Financial funds buying shares on the open market (reporting threshold crossings) are NOT strategic M&A deals.
- Reject "Insider Trades": CEO/Director buying company stock.
- Reject awards, contracts, earnings, financings, product launches, and market reports unless a government tender/auction explicitly grants asset transfer (lease, license, concession, block, acreage, or project rights).

ACTIVE MARKET CONTEXT (USE THIS FOR CROSS-ARTICLE INTELLIGENCE):
{run_context}
"""


class AsyncAIExtractor:
    """ARIA-powered multi-model, multi-key Groq AI extractor with Chain-of-Thought reasoning."""

    def __init__(self, api_key: str | None = None, extra_keys: list[str] | None = None):
        # Build key pool: explicit key(s) + env vars
        self._all_keys: list[str] = []
        if api_key:
            self._all_keys.append(api_key)
        if extra_keys:
            self._all_keys.extend(extra_keys)
        # Add any env-var keys not already present
        for k in _collect_groq_keys():
            if k not in self._all_keys:
                self._all_keys.append(k)
        self.groq_key = self._all_keys[0] if self._all_keys else None
        self.enabled = bool(self.groq_key)

        # ── Groq Client Pool (one client per key) ──
        self._groq_clients: list = []
        self._client_index = 0
        self._model_index = 0
        self._model_index_lock = asyncio.Lock()

        # Track failures per model to skip broken ones
        self._model_failures = {m: 0 for m in _GROQ_MODELS}
        # Space out calls to avoid org-level generic limits
        self._last_call_time = 0.0
        self._global_lock = asyncio.Lock()

        if self.enabled:
            try:
                from groq import AsyncGroq
                for key in self._all_keys:
                    self._groq_clients.append(AsyncGroq(api_key=key, max_retries=0))
                logger.info(f"Groq pool: {len(self._groq_clients)} key(s) × {len(_GROQ_MODELS)} models = "
                           f"{len(self._groq_clients) * len(_GROQ_MODELS)} rotation slots")
            except ImportError:
                logger.error("Groq SDK not installed! Run: pip install groq")
                self.enabled = False

        # ── Self-Learning Memory ──
        self._known_companies: set[str] = set()
        self._known_deals: list[dict] = []
        self._corrections: dict = {
            "wrong_accepts": [], "wrong_rejects": [],
            "keywords_accept": [], "keywords_reject": []
        }
        self._load_memory()
        self._load_feedback_from_excel()

        # Budget mode controls prompt size and optional AI usage.
        self.budget_mode = _AI_BUDGET_MODE
        self._verify_body_chars = 800 if self.budget_mode else 1200
        self._extract_body_chars = 1500 if self.budget_mode else 3000
        self._precedent_limit = 1 if self.budget_mode else 3
        self._verify_body_chars_rich = 1600 if self.budget_mode else self._verify_body_chars
        self._extract_body_chars_rich = 3500 if self.budget_mode else self._extract_body_chars

        # ── System Prompt ──
        self.system_instruction = (
            "You are an expert energy-sector M&A data extractor. "
            "Extract structured entity data only from articles describing a firm HIGH-VALUE CORPORATE TRANSACTION. "
            "TRANSACTION DEFINITION: Acquisition, Merger, Stake Sale, JV Formation, Asset Purchase, Spin-off, Portfolio optimization, OR Major Government Tender/Auction Wins (e.g., Renewable Energy capacity auctions).\n\n"
            "STRICT REJECTION CATEGORIES (MUST REJECT):\n"
            "- NON-ENERGY SECTORS: Semiconductors, Power Chips, Microchips, Crypto/Bitcoin Mining, HVAC, GPU/Computing Infrastructure, Cement, Pharmaceuticals/Biotech, Retail, Fashion, Sports, Media.\n"
            "- OPERATIONAL NEWS: Well drilling, production updates, field results, CEO hires, board changes.\n"
            "- MARKET REPORTS: Market Size, CAGR, 'Market to reach $X', 'Top players in...'.\n"
            "- AWARDS/FINANCINGS/CONTRACTS/PRODUCT LAUNCHES: Reject awards, financing packages, earnings, contracts, product launches, and market reports unless a GOVERNMENT tender/auction explicitly grants asset transfer (lease, license, concession, block, acreage, or project rights).\n"
            "- INSTITUTIONAL PORTFOLIO TRADES: When a FINANCIAL institution (fund, bank, investment advisor, asset manager, ETF) "
            "merely buys or sells publicly listed SHARES on the open market, even in an energy company — this is NOT a strategic deal. "
            "Key signals to reject: 'investment advisors', 'fund', 'LLC increases stake', 'boosts stake', 'reduces position', 'cuts stake', 'crosses threshold'. "
            "EXCEPTION: If an Asset Manager/Fund buys a significant or strategic stake directly in a private infrastructure asset (e.g. 'AllianzGI takes stake in German grid operator Amprion'), this IS a valid strategic deal.\n\n"
            "Normalization Rules:\n"
            "1. Remove corporate suffixes ('Inc', 'LLC', etc.).\n"
            "2. 'Exxon Mobil' -> 'ExxonMobil'.\n"
            "3. 'BP plc' -> 'BP'.\n"
            "Output STRICTLY as JSON with these exact keys: buyer, seller, asset, value, industry, sector."
        )

        # ── Intelligence Systems ──
        self.context_manager = RunContextManager()
        self.persona = ARIA_SYSTEM_PROMPT

        # Stats
        self._stats = {
            "calls": 0, "failures": 0, "rate_limits_429": 0,
            "auto_approves": 0, "corrections_applied": 0,
            "fail_safe_reviews": 0,
        }

        if self.enabled:
            logger.info(f"AI Extractor enabled — Provider: Groq (Multi-Model Rotation)")

    # ──────────────────────────────────────────────────
    # Memory & Learning System
    # ──────────────────────────────────────────────────

    def _load_memory(self):
        try:
            if _MEMORY_FILE.exists():
                data = json.loads(_MEMORY_FILE.read_text(encoding='utf-8'))
                self._known_companies = set(data.get("companies", []))
                self._known_deals = data.get("deals", [])[-500:] # Increased memory
                self._corrections = data.get("corrections", self._corrections)
                logger.info(f"Loaded {len(self._known_companies)} known companies, "
                          f"{len(self._corrections.get('wrong_rejects', []))} rejection corrections")
        except Exception as e:
            logger.warning(f"Could not load deal memory: {e}")

    def load_shared_feedback(self, feedback_list: list[str]):
        """Load collaborative feedback from Google Sheets to inject into AI prompt."""
        self._shared_feedback = feedback_list
        if feedback_list:
            logger.info(f"Loaded {len(feedback_list)} shared feedback entries into AI context.")

    def save_memory(self):
        try:
            data = {
                "companies": sorted(self._known_companies),
                "deals": self._known_deals[-200:],
                "corrections": self._corrections,
                "last_updated": datetime.now().isoformat(),
                "stats": self._stats,
            }
            _MEMORY_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
            logger.info(f"✅ Saved {len(self._known_companies)} companies to memory")
        except Exception as e:
            logger.warning(f"Could not save deal memory: {e}")

    def learn_from_deal(self, deal: dict):
        for field in ["buyer", "seller", "Buyer", "Seller"]:
            name = deal.get(field, "Unknown")
            if name and name != "Unknown" and len(name) > 2:
                clean = name.strip().rstrip("., ").lower()
                self._known_companies.add(clean)
        self._known_deals.append({
            "headline": deal.get("Headline", deal.get("headline", ""))[:100],
            "buyer": deal.get("Buyer", deal.get("buyer", "Unknown")),
            "seller": deal.get("Seller", deal.get("seller", "Unknown")),
        })

    def _load_feedback_from_excel(self):
        try:
            import openpyxl
        except ImportError:
            return

        search_dirs = [RUN_REPORTS_DIR, ARCHIVE_DIR, _MEMORY_DIR]
        excel_files = []
        for base_dir in search_dirs:
            excel_files.extend(glob.glob(str(base_dir / "Energy_MA_Report_*.xlsx")))
        if not excel_files:
            return

        latest = max(excel_files, key=os.path.getmtime)
        logger.info(f"Checking feedback from: {Path(latest).name}")

        try:
            wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
        except PermissionError:
            logger.warning("Excel file is open/locked — skipping feedback read")
            return
        except Exception as e:
            logger.warning(f"Could not read Excel for feedback: {e}")
            return

        corrections_found = 0

        if "Output" in wb.sheetnames:
            ws = wb["Output"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            ai_col = headline_col = feedback_col = None
            for i, h in enumerate(headers):
                if h and "AI Correct" in str(h): ai_col = i
                if h and h == "Headline": headline_col = i
                if h and "Feedback to AI" in str(h): feedback_col = i

            if ai_col is not None and headline_col is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > max(ai_col, headline_col):
                        feedback_yn = str(row[ai_col] or "").strip().lower()
                        headline = str(row[headline_col] or "")
                        # Read explanation from 'Feedback to AI' column if present
                        explanation = ""
                        if feedback_col is not None and len(row) > feedback_col:
                            explanation = str(row[feedback_col] or "").strip()
                        if feedback_yn == "no" and headline:
                            entry = headline[:150]
                            if explanation:
                                # Store as "HEADLINE ||| EXPLANATION" for richer AI prompting
                                entry = f"{headline[:120]} ||| {explanation[:200]}"
                            self._corrections["wrong_accepts"].append(entry)
                            corrections_found += 1

        if "Rejected by AI" in wb.sheetnames:
            ws = wb["Rejected by AI"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            ai_col = headline_col = feedback_col = None
            for i, h in enumerate(headers):
                if h and "AI Correct" in str(h): ai_col = i
                if h and h == "Headline": headline_col = i
                if h and "Feedback to AI" in str(h): feedback_col = i

            if ai_col is not None and headline_col is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > max(ai_col, headline_col):
                        feedback_yn = str(row[ai_col] or "").strip().lower()
                        headline = str(row[headline_col] or "")
                        explanation = ""
                        if feedback_col is not None and len(row) > feedback_col:
                            explanation = str(row[feedback_col] or "").strip()
                        if feedback_yn == "no" and headline:
                            entry = headline[:150]
                            if explanation:
                                entry = f"{headline[:120]} ||| {explanation[:200]}"
                            self._corrections["wrong_rejects"].append(entry)
                            for word in headline.split():
                                if len(word) > 4 and word[0].isupper():
                                    self._known_companies.add(word.lower())
                            corrections_found += 1

        if "User Feedback for AI Learning" in wb.sheetnames:
            ws = wb["User Feedback for AI Learning"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    feedback_text = str(row[0]).strip()
                    if feedback_text:
                        if any(kw in feedback_text.lower() for kw in ["accept", "include", "should be", "is a deal"]):
                            self._corrections["keywords_accept"].append(feedback_text[:200])
                        elif any(kw in feedback_text.lower() for kw in ["reject", "exclude", "not a deal", "wrong"]):
                            self._corrections["keywords_reject"].append(feedback_text[:200])
                        corrections_found += 1

        wb.close()
        if corrections_found:
            logger.info(f"📚 Loaded {corrections_found} user corrections")
            self._stats["corrections_applied"] += corrections_found

    def _check_known_company(self, headline: str) -> str | None:
        if not self._known_companies:
            return None
        h_lower = headline.lower()
        for company in self._known_companies:
            if len(company) > 3 and company in h_lower:
                return company
        return None

    def _check_corrections(self, headline: str) -> dict | None:
        h_lower = headline.lower()

        for wr in self._corrections.get("wrong_rejects", []):
            wr_words = set(wr.lower().split())
            h_words = set(h_lower.split())
            overlap = len(wr_words & h_words)
            if overlap >= 3 or (overlap >= 2 and len(wr_words) <= 5):
                self._stats["corrections_applied"] += 1
                return {"is_deal": True, "reason": "Correction override: similar to user-corrected headline"}

        for kw in self._corrections.get("keywords_accept", []):
            if any(w in h_lower for w in kw.lower().split() if len(w) > 4):
                return {"is_deal": True, "reason": "Correction keyword match: user feedback"}

        return None

    # ──────────────────────────────────────────────────
    # Groq Multi-Model API Layer
    # ──────────────────────────────────────────────────

    async def _get_next_model(self) -> tuple[object, str] | None:
        """Round-robin across (client, model) pairs."""
        async with self._model_index_lock:
            n_clients = max(1, len(self._groq_clients))
            total_slots = n_clients * len(_GROQ_MODELS)
            for _ in range(total_slots):
                client_idx = self._client_index % n_clients
                model = _GROQ_MODELS[self._model_index % len(_GROQ_MODELS)]
                # Advance both indexes so next call uses a different combo
                self._model_index += 1
                if self._model_index % len(_GROQ_MODELS) == 0:
                    self._client_index += 1
                if self._model_failures.get(model, 0) < 5:
                    return self._groq_clients[client_idx], model
        return None

    async def _call_groq(
        self,
        messages: list[dict],
        use_json: bool = True,
        json_mode: str = "object",
    ) -> dict | list | str | None:
        if not self._groq_clients:
            return None

        next_slot = await self._get_next_model()
        if not next_slot:
            return None
        client, model = next_slot

        # Per-call spacing: 0.5s between calls (safe with 4+ keys × 3 models)
        async with self._global_lock:
            elapsed = time.time() - self._last_call_time
            if elapsed < 0.5:
                await asyncio.sleep(0.5 - elapsed)
            self._last_call_time = time.time()

        kwargs = {
            "messages": messages,
            "model": model,
            "temperature": 0.0,
        }
        if use_json and json_mode == "object":
            kwargs["response_format"] = {"type": "json_object"}

        try:
            # Task-Level Timeout: Groq calls should NEVER take 45+ seconds.
            # If the network or provider hangs, time out and failover to the next model.
            response = await asyncio.wait_for(client.chat.completions.create(**kwargs), timeout=45)
            self._model_failures[model] = 0
            self._stats["calls"] += 1
            raw = str(response.choices[0].message.content or "")
            
            if not use_json:
                return raw
                
            parsed = self._greedy_json_parse(raw)
            if parsed is not None:
                return parsed
            
            logger.warning(f"Groq returned invalid JSON on [{model}]. Raw text: {raw[:100]}")
            # Instead of returning None, return the raw text so _generate can try to repair it
            return {"_raw_unparsed": raw}
            
        except Exception as e:
            err_str = str(e)
            if "429" in err_str:
                self._stats["rate_limits_429"] += 1
                self._model_failures[model] = self._model_failures.get(model, 0) + 1
                logger.warning(f"Groq 429 on model [{model}] (failure #{self._model_failures[model]}). Rotating immediately.")
                # Instead of sleeping, return a special code to trigger immediate rotation in _generate
                return {"_retry_immediately": True}
            else:
                self._model_failures[model] = self._model_failures.get(model, 0) + 1
                logger.warning(f"Groq error on [{model}]: {err_str[:150]}")
            
            await asyncio.sleep(2)
            return None

    async def _generate(
        self,
        messages: list[dict],
        use_json: bool = True,
        json_mode: str = "object",
    ) -> dict | list | str | None:
        """Multi-model generation with automatic failover and JSON self-repair."""
        # Try up to 8 different models/keys on failure due to 429 limits
        for attempt in range(min(8, len(_GROQ_MODELS) * max(1, len(self._groq_clients)))):
            result = await self._call_groq(messages, use_json, json_mode=json_mode)
            if result is not None:
                if isinstance(result, dict) and result.get("_retry_immediately"):
                    continue # Try the next key/model combo instantly
                
                # If we didn't want JSON, just return the raw string
                if not use_json and isinstance(result, str):
                    return result

                # Check if we have unparsed raw text that needs repair
                if isinstance(result, dict) and "_raw_unparsed" in result:
                    if not use_json: # Should not happen with new _call_groq but defensive
                        return result["_raw_unparsed"]
                        
                    raw_text = result["_raw_unparsed"]
                    # TRIGGER SELF-REPAIR: Send the raw text to a model to fix the JSON
                    repair_messages = [
                        {"role": "system", "content": "You are a JSON repair bot. Your ONLY job is to take malformed text and turn it into a valid JSON object. No prose. No intro."},
                        {"role": "user", "content": f"Fix this malformed JSON to be valid and follow the original structure:\n\n{raw_text}"}
                    ]
                    # Attempt a single repair call (non-recursive)
                    repaired = await self._call_groq(repair_messages, use_json=True, json_mode=json_mode)
                    if isinstance(repaired, (dict, list)) and not (
                        isinstance(repaired, dict) and "_raw_unparsed" in repaired
                    ):
                        logger.info("Successfully repaired malformed JSON")
                        return repaired
                    return None # Repair failed
                
                return result
        return None

    def _greedy_json_parse(self, text: str) -> dict | list | None:
        """Extract the first valid JSON payload from a potentially conversational string."""
        cleaned = (text or "").strip()
        if not cleaned:
            return None

        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
            cleaned = re.sub(r"\s*```$", "", cleaned)

        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            pass

        start = cleaned.find('[')
        end = cleaned.rfind(']')
        if start != -1 and end != -1 and end > start:
            inner = cleaned[start:end+1]
            try:
                return json.loads(inner.strip())
            except json.JSONDecodeError:
                pass

        start = cleaned.find('{')
        end = cleaned.rfind('}')
        if start != -1 and end != -1 and end > start:
            inner = cleaned[start:end+1]
            try:
                return json.loads(inner.strip())
            except json.JSONDecodeError:
                pass
        return None

    def early_reject_reason(self, headline: str, body_snippet: str = "") -> str | None:
        """
        Conservative, zero-cost hard reject gate to prevent obvious non-deals
        from ever hitting the AI call path. Returns a short reason or None.
        """
        headline_lower = (headline or "").lower()
        body_lower = (body_snippet or "").lower()
        combined_text = f"{headline_lower} {body_lower}"

        # If any strong M&A verb appears, do NOT early reject.
        ma_verbs = [
            "acquire", "acquires", "acquired", "acquisition", "merge", "merger",
            "merged", "divest", "divestiture", "buy", "buys", "purchase",
            "sell", "sale", "stake", "joint venture", "jv", "mou", "loi",
            "asset sale", "asset purchase", "farm-in", "farm-out",
        ]
        if any(v in combined_text for v in ma_verbs):
            return None

        def _has_any(terms: list[str], text: str) -> bool:
            return any(t in text for t in terms)

        gov_terms = [
            "government", "ministry", "department", "regulator", "commission",
            "federal", "state", "province", "municipal", "public authority",
            "tender", "auction", "bid round", "concession", "license", "permit"
        ]
        award_terms = [
            "award", "awarded", "wins award", "tender award", "contract award",
            "bid award", "concession award"
        ]
        transfer_terms = [
            "lease", "license", "concession", "block", "acreage",
            "asset", "project rights", "ownership", "transfer", "rights"
        ]
        is_gov_award_transfer = (
            _has_any(gov_terms, combined_text)
            and _has_any(award_terms, combined_text)
            and _has_any(transfer_terms, combined_text)
        )

        award_reject_patterns = [
            "award", "awarded", "awards", "best practices award",
            "industry award", "recognition", "honor", "prize", "trophy"
        ]
        financing_reject_patterns = [
            "financing", "financings", "financing package", "credit facility",
            "debt facility", "loan facility", "notes offering", "bond offering",
            "private placement", "equity offering", "at-the-market", "atm program",
            "capital raise", "raises $", "raised $", "secures financing"
        ]
        earnings_reject_patterns = [
            "earnings", "financial results", "quarterly results", "q1 results",
            "q2 results", "q3 results", "q4 results", "fiscal year", "guidance",
            "eps", "ebitda", "revenue", "profit"
        ]
        contract_reject_patterns = [
            "contract award", "awarded contract", "wins contract",
            "service contract", "supply agreement", "framework agreement",
            "purchase order", "service agreement", "task order"
        ]
        product_reject_patterns = [
            "launches", "launch", "unveils", "debuts", "introduces",
            "new product", "product line", "solution launch"
        ]
        market_report_patterns = [
            "market report", "market size", "cagr", "market to reach",
            "market forecast", "market analysis", "research report", "industry report"
        ]

        if _has_any(market_report_patterns, combined_text):
            return "Hard reject: market report"
        if _has_any(earnings_reject_patterns, combined_text):
            return "Hard reject: earnings/results"
        if _has_any(financing_reject_patterns, combined_text):
            return "Hard reject: financing package"
        if _has_any(product_reject_patterns, combined_text):
            return "Hard reject: product launch"
        if _has_any(contract_reject_patterns, combined_text) and not is_gov_award_transfer:
            return "Hard reject: contract/award"
        if _has_any(award_reject_patterns, combined_text) and not is_gov_award_transfer:
            return "Hard reject: award/recognition"

        return None

    # ──────────────────────────────────────────────────
    # AI Deal Verification Gate
    # ──────────────────────────────────────────────────

    async def verify_is_deal(self, headline: str, body_snippet: str = "") -> dict:
        """
        ARIA-powered verification gate. Uses Chain-of-Thought reasoning to confirm
        if a headline describes a strategic M&A or structural transaction.
        """
        headline_lower = headline.lower()
        body_lower = (body_snippet or "").lower()
        combined_text = f"{headline_lower} {body_lower}"

        def _has_any(terms: list[str], text: str) -> bool:
            return any(t in text for t in terms)

        # Allow government awards only when they grant asset transfer rights.
        gov_terms = [
            "government", "ministry", "department", "regulator", "commission",
            "federal", "state", "province", "municipal", "public authority",
            "tender", "auction", "bid round", "concession", "license", "permit"
        ]
        award_terms = [
            "award", "awarded", "wins award", "tender award", "contract award",
            "bid award", "concession award"
        ]
        transfer_terms = [
            "lease", "license", "concession", "block", "acreage",
            "asset", "project rights", "ownership", "transfer", "rights"
        ]
        is_gov_award_transfer = (
            _has_any(gov_terms, combined_text)
            and _has_any(award_terms, combined_text)
            and _has_any(transfer_terms, combined_text)
        )

        # ── Tier 0: Instant Pattern Rejection (Zero AI Cost) ──
        # These patterns fire BEFORE any LLM call.
        # Adding patterns here is the single highest ROI cost reduction measure.
        # Organised into 8 semantic categories for maintainability.
        award_reject_patterns = [
            "award", "awarded", "awards", "best practices award",
            "industry award", "recognition", "honor", "prize", "trophy"
        ]
        financing_reject_patterns = [
            "financing", "financings", "financing package", "credit facility",
            "debt facility", "loan facility", "notes offering", "bond offering",
            "private placement", "equity offering", "at-the-market", "atm program",
            "capital raise", "raises $", "raised $", "secures financing"
        ]
        earnings_reject_patterns = [
            "earnings", "financial results", "quarterly results", "q1 results",
            "q2 results", "q3 results", "q4 results", "fiscal year", "guidance",
            "eps", "ebitda", "revenue", "profit"
        ]
        contract_reject_patterns = [
            "contract award", "awarded contract", "wins contract",
            "service contract", "supply agreement", "framework agreement",
            "purchase order", "service agreement", "task order"
        ]
        product_reject_patterns = [
            "launches", "launch", "unveils", "debuts", "introduces",
            "new product", "product line", "solution launch"
        ]
        market_report_patterns = [
            "market report", "market size", "cagr", "market to reach",
            "market forecast", "market analysis", "research report", "industry report"
        ]

        for pat in award_reject_patterns:
            if pat in headline_lower and not is_gov_award_transfer:
                logger.info(f"Instant Reject (Award): {headline[:60]}")
                return {"is_deal": False, "reason": f"Auto-rejected: award pattern '{pat}'"}
        for pat in financing_reject_patterns:
            if pat in headline_lower:
                logger.info(f"Instant Reject (Financing): {headline[:60]}")
                return {"is_deal": False, "reason": f"Auto-rejected: financing pattern '{pat}'"}
        for pat in earnings_reject_patterns:
            if pat in headline_lower:
                logger.info(f"Instant Reject (Earnings): {headline[:60]}")
                return {"is_deal": False, "reason": f"Auto-rejected: earnings pattern '{pat}'"}
        for pat in contract_reject_patterns:
            if pat in headline_lower and not is_gov_award_transfer:
                logger.info(f"Instant Reject (Contract): {headline[:60]}")
                return {"is_deal": False, "reason": f"Auto-rejected: contract pattern '{pat}'"}
        for pat in product_reject_patterns:
            if pat in headline_lower:
                logger.info(f"Instant Reject (Product): {headline[:60]}")
                return {"is_deal": False, "reason": f"Auto-rejected: product pattern '{pat}'"}
        for pat in market_report_patterns:
            if pat in headline_lower:
                logger.info(f"Instant Reject (Market Report): {headline[:60]}")
                return {"is_deal": False, "reason": f"Auto-rejected: market report pattern '{pat}'"}

        instant_reject_patterns = [

            # ── Category A: Insider / Director stock trades (NOT M&A) ──
            "insider mark", "insider sell", "insider buys",
            "sells shares", "sells stock", "buys shares",
            "officer sells", "director sells", "director buys",
            "shares of", "general interest",
            # ── Category B: Institutional open-market portfolio trades ──
            "investment advisors acquires", "investment advisors increases",
            "investment advisors reduces", "investment advisors cuts",
            "asset management acquires", "asset management increases",
            "asset management reduces", "asset management cuts",
            "investment board buys", "mutual fund picks up",
            "boosts stake in", "raises stake in", "lifts stake in",
            "reduces stake in", "cuts stake in", "trims stake in",
            "crosses reporting threshold", "reporting threshold",
            "new position in", "increases position in", "decreases position in",
            "buys rs ", "buys ₹",
            "llc buys shares", "llc reduces", "fund acquires shares",
            # ── Category C: Financial noise / rates ──
            "savings interest rate", "mortgage rate", "refinance rate",
            "high-yield savings", "credit card rates", "home equity loan",
            "heloc rate", "cd rate", "personal loan", "auto loan rate",
            # ── Category D: Commodity price / market moves (NOT deals) ──
            "as april contract expires", "commodity gains", "commodity prices",
            "natural gas gains", "oil soars", "oil rallies", "oil tumbles",
            "crude oil falls", "crude rises", "brent crude", "wti crude",
            "oil prices today", "gas prices today", "price of oil",
            "market rally", "stocks rise", "stocks fall",
            # ── Category E: Service contracts (NOT acquisitions) ──
            "wins service contract", "awarded service contract",
            "wins drilling contract", "awarded drilling contract",
            "secures maintenance contract", "secures service agreement",
            "wins maintenance contract", "awarded maintenance contract",
            "task order", "framework agreement", "supply agreement",
            # ── Category I: Regulatory / policy updates ──
            "epa approves", "nationwide sales of e15", "removes barriers to sale",
            "final rule for", "policy update",
            # ── Category F: Market commentary / analyst opinion ──
            "i want to buy", "no-brainer stock", "stock to buy",
            "stocks to watch", "best stocks", "top stocks",
            "jim cramer", "score card", "scorecard", "analyst rating",
            "price target", "buy rating", "sell rating", "hold rating",
            "earnings preview", "earnings estimate",
            # ── Category G: SPAC stub / index pages ──
            "about acquisition corp", "about spac",
            "quarterly balance sheet", "earnings per share", "eps guidance",
            "on earnings, business", "reports revenue",
            "reports quarterly", "quarterly results",
            # ── Category H: Hard off-sector rejections ──
            # Pharma / Biotech
            "clinical trial", "drug pipeline", "fda approval",
            "therapeutic", "pharma acquisition" if "energy" not in headline_lower else "",
            "biotech", "biopharmaceutical",
            # HVAC / Appliances / Retail
            "appliance repair", "hvac", "air conditioning unit",
            "cosmetics", "skincare", "home depot", "mingledorff",
            "wine & spirits", "wine and spirits",
            # Semiconductors / Tech
            "power chip", "semiconductor", "microchip",
            "gpu cluster", "data center chips", "artificial intelligence chip",
            # Crypto / Mining (non-energy)
            "bitcoin mining", "crypto mining", "cryptocurrency",
            "digital assets fund",
            # Sports / Entertainment
            "nfl", "nba", "premier league", "football club",
            "entertainment acquisition" if "energy" not in headline_lower else "",
        ]
        for pat in instant_reject_patterns:
            if pat and pat in headline_lower:
                logger.info(f"🚫 Instant Reject (Pattern): {headline[:60]}")
                return {"is_deal": False, "reason": f"Auto-rejected: matches pattern '{pat}'"}

        # Clean body snippet
        body_snippet = self._clean_text(body_snippet)

        # ── Tier 1: Correction overrides from user feedback (highest priority) ──
        correction = self._check_corrections(headline)
        if correction:
            logger.info(f"🟢 Correction override: {headline[:60]}")
            return correction

        if not self.enabled:
            self._stats["fail_safe_reviews"] += 1
            return {
                "is_deal": True,
                "reason": "AI verification unavailable — sent to Review Queue for manual check",
                "_fail_safe": True,
            }

        # ── Tier 2: Prepare ARIA Persona with Live Context ──
        if self.budget_mode:
            system_prompt = (
                "You are an energy/mining M&A verification analyst. "
                "Accept ONLY real corporate transactions (acquisition, merger, stake sale, JV, asset sale/purchase) "
                "or government tenders/auctions that grant asset transfer rights (lease/license/concession/block/acreage/project rights). "
                "Reject awards, financing packages, earnings, contracts, product launches, market reports, "
                "operational updates, portfolio trades, insider trades, and non-energy sectors."
            )
        else:
            run_context = self.context_manager.get_context_summary()
            system_prompt = self.persona.format(run_context=run_context)

        def _build_user_content(body_chars: int) -> str:
            content = f"HEADLINE: {headline}"
            if body_snippet:
                content += f"\n\nARTICLE BODY SNIPPET (first {body_chars} chars):\n{body_snippet[:body_chars]}"

            known = self._check_known_company(headline)
            if known:
                content += f"\n\n[ANALYST HINT]: '{known}' is a known energy sector entity in our database. Verify if this news is a TRANSACTION involving them."

            learned_rules = ""
            if self._corrections.get("wrong_rejects"):
                examples = self._corrections["wrong_rejects"][-self._precedent_limit:]
                learned_rules += "\n\nCRITICAL PRECEDENTS (USER FEEDBACK - MUST ACCEPT):\n"
                for ex in examples:
                    learned_rules += f"- {ex[:120]}\n"
            if self._corrections.get("wrong_accepts"):
                examples = self._corrections["wrong_accepts"][-self._precedent_limit:]
                learned_rules += "\nCRITICAL PRECEDENTS (USER FEEDBACK - MUST REJECT):\n"
                for ex in examples:
                    learned_rules += f"- {ex[:120]}\n"

            if learned_rules:
                content += learned_rules

            content += "\n\nRESPONSE FORMAT: Output JSON with these keys:\n"
            content += "  is_deal: true/false\n"
            content += "  confidence: 0.0–1.0\n"
            content += "  reasoning: 1-2 sentences on why it is or is not a deal\n"
            content += "  reject_reason: if not a deal, brief reason\n"
            content += "  sector_id: primary sector tag\n"
            content += "  sheet: ONLY if is_deal=true, assign ONE of [Upstream, Midstream, OFS, R&M, P&U, JV & Partnerships].\n"
            content += "\nSHEET GUIDE (assign based on the TARGET asset/company):\n"
            content += "  Upstream = oil/gas E&P, fields, wells, reserves, acreage, basins, farm-in/farm-out\n"
            content += "  Midstream = pipelines, LNG terminals, gas processing, fractionators, NGL infrastructure\n"
            content += "  OFS = oilfield services, drilling contractors, rig companies, hydraulic fracturing, subsea equipment\n"
            content += "  R&M = refineries, petrochemicals, fuel retail, lubricants, crackers\n"
            content += "  P&U = power/utilities, solar, wind, hydro, nuclear, battery storage, hydrogen, mining (gold/copper/lithium), clean energy\n"
            content += "  JV & Partnerships = joint ventures, partnerships, MOUs, LOIs, strategic alliances\n"
            content += "  REJECT = NOT energy/mining/power sector at all (hotel, pharma, consumer goods, social media, etc.)\n"
            return content

        user_content = _build_user_content(self._verify_body_chars)
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content}
        ]

        # ── Tier 3: AI Inference ──
        verdict = await self._safe_generate(messages)

        if verdict.get("_fail_safe"):
            return verdict

        is_deal = verdict.get("is_deal", False)
        reasoning = verdict.get("reasoning", "No reasoning provided")
        reject_reason = verdict.get("reject_reason", "n/a")
        confidence = verdict.get("confidence", 0.0)

        log_msg = f"🧠 ARIA Verdict for: {headline[:60]} -> {'DEAL' if is_deal else 'REJECT'}"
        if is_deal:
            logger.info(f"{log_msg} (Conf: {confidence:.2f})")
        else:
            logger.info(f"{log_msg} (Reason: {reject_reason})")
        
        # ── Tier 3.5: Escalate ambiguous cases to richer prompt (budget mode only) ──
        if self.budget_mode and not verdict.get("_fail_safe"):
            strict_reasons = [
                "pharma", "biotech", "chip", "semiconductor", "crypto", "bitcoin",
                "hvac", "market report", "personnel", "insider", "commentary",
                "award", "financing", "contract", "earnings", "product launch"
            ]
            reason_lower = reject_reason.lower()
            is_strict_reject = any(x in reason_lower for x in strict_reasons)
            should_escalate = (
                (0.35 <= confidence <= 0.75) or
                (not is_deal and not is_strict_reject)
            )
            if should_escalate and body_snippet:
                run_context = self.context_manager.get_context_summary()
                rich_system = self.persona.format(run_context=run_context)
                rich_user = _build_user_content(self._verify_body_chars_rich)
                rich_messages = [
                    {"role": "system", "content": rich_system},
                    {"role": "user", "content": rich_user}
                ]
                rich_verdict = await self._safe_generate(rich_messages)
                if rich_verdict and not rich_verdict.get("_fail_safe"):
                    verdict = rich_verdict
                    is_deal = verdict.get("is_deal", False)
                    reasoning = verdict.get("reasoning", reasoning)
                    reject_reason = verdict.get("reject_reason", reject_reason)
                    confidence = verdict.get("confidence", confidence)

        # ── Tier 4: Cross-Check Soft Rejections ──
        if is_deal:
            # Use AI sheet if returned, else leave empty for keyword fallback in scraper
            ai_sheet = verdict.get("sheet", "").strip()
            # If AI says REJECT in the sheet field, treat as rejected even if is_deal=True
            if ai_sheet == "REJECT":
                return {"is_deal": False, "reason": "AI classified as non-energy sector in sheet field"}
            return {"is_deal": True, "reason": reasoning, "confidence": confidence, "sheet": ai_sheet}

        reason_lower = reject_reason.lower()
        strict_rejects = [
            "pharma", "biotech", "chip", "semiconductor", "crypto", "bitcoin",
            "hvac", "market report", "personnel", "insider", "commentary",
            "award", "financing", "contract", "earnings", "product launch"
        ]
        is_strict = any(x in reason_lower for x in strict_rejects)

        if not is_strict:
            logger.info(f"🔍 Perform cross-check for soft rejection: {headline[:60]}")
            cross_check_confirmed = await self._google_news_cross_check(headline)
            if cross_check_confirmed:
                return {"is_deal": True, "reason": "Initially rejected by AI, but confirmed via Google News cross-check.", "confidence": 0.6}

        return {"is_deal": False, "reason": reject_reason}

    async def _safe_generate(self, messages: list[dict]) -> dict:
        for attempt in range(3):
            result = await self._generate(messages, use_json=True)
            if result is not None:
                # Guard: AI sometimes returns a JSON array — unwrap if so
                if isinstance(result, list):
                    result = result[0] if result else {}
                if not isinstance(result, dict):
                    if attempt < 2:
                        await asyncio.sleep(2 + attempt * 2)
                    continue
                is_deal = result.get("is_deal")
                if isinstance(is_deal, str):
                    is_deal = is_deal.lower() in ("true", "1", "yes")
                return {
                    "is_deal": bool(is_deal),
                    "reason": str(result.get("reason", result.get("reasoning", ""))),
                    "reasoning": str(result.get("reasoning", "")),
                    "reject_reason": str(result.get("reject_reason", "n/a")),
                    "confidence": float(result.get("confidence", 0.0) or 0.0),
                    "sheet": str(result.get("sheet", "")).strip(),
                }
            if attempt < 2:
                await asyncio.sleep(2 + attempt * 2)

        self._stats["fail_safe_reviews"] += 1
        logger.warning(f"AI FAIL-SAFE: All model calls failed → sending to Review Queue")
        return {
            "is_deal": True,
            "reason": "AI verification unavailable — sent to Review Queue for manual check",
            "_fail_safe": True,
        }

    async def _google_news_cross_check(self, headline: str) -> bool:
        try:
            import feedparser
            from urllib.parse import quote_plus

            words = headline.split()[:6]
            search_query = " ".join(words) + " acquisition OR deal OR merger OR buys OR sells"
            encoded = quote_plus(search_query)
            rss_url = f"https://news.google.com/rss/search?q={encoded}&hl=en-US&gl=US&ceid=US:en"

            def _sync_check():
                feed = feedparser.parse(rss_url)
                return len(feed.entries)

            count = await asyncio.to_thread(_sync_check)
            if count >= 5:
                logger.info(f"Cross-check CONFIRMED: {count} results for '{headline[:50]}'")
                return True
            else:
                logger.info(f"Cross-check negative: {count} results for '{headline[:50]}'")
                return False

        except Exception as e:
            logger.warning(f"Cross-check failed: {e}")
            return False

    async def translate_full_body(self, text: str, source_lang: str = "Russian") -> str:
        """Translates a full article body into English, preserving technical M&A context."""
        if not self.enabled or not text:
            return text

        messages = [
            {
                "role": "system",
                "content": (
                    f"You are a professional technical translator specializing in the {source_lang} energy sector. "
                    "Translate the provided text into professional English. Maintain technical terms (e.g., 'farm-in', 'JV', 'wellhead'). "
                    "Return ONLY the translated text."
                )
            },
            {"role": "user", "content": f"Translate the following {source_lang} article body to English:\n\n{text[:5000]}"}
        ]
        
        # Use robust _generate with failover for translations
        result = await self._generate(messages, use_json=False)
        if result and isinstance(result, str):
            return result
        return text

    # ──────────────────────────────────────────────────
    # Entity Extraction
    # ──────────────────────────────────────────────────

    def _clean_text(self, text: str) -> str:
        """Strip SEC boilerplate, standard headers, and excessive noise before AI processing."""
        if not text: return ""
        # Remove standard SEC headers
        noise_patterns = [
            r"Filed by the Registrant", r"SEC\.gov", r"United States Securities and Exchange Commission",
            r"Check the appropriate box", r"Pursuant to Section", r"Form S-1", r"Form S-4",
            r"Table of Contents", r"Exhibit \d+", r"\[X\]", r"\[ \]"
        ]
        cleaned = text
        for pat in noise_patterns:
            cleaned = re.sub(pat, "", cleaned, flags=re.IGNORECASE)
        # Normalize whitespace
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        return cleaned

    async def extract_deal(self, headline: str, body_text: str) -> dict | None:
        """
        Extracts structured deal data using the ARIA persona and live context.
        Enriches the deal with strategic rationale and geographic metadata.
        """
        if not self.enabled:
            return None
            
        body_text = self._clean_text(body_text)
        
        # Prepare ARIA Persona with Live Context (budget mode uses compact prompt)
        if self.budget_mode:
            system_prompt = (
                "You are an energy/mining M&A data extractor. "
                "Extract structured entity data ONLY from real transactions: acquisition, merger, stake sale, JV, "
                "asset purchase/sale, or government tender/auction granting asset transfer (lease/license/concession/block/acreage/project rights). "
                "Reject awards, financing packages, earnings, contracts, product launches, and market reports. "
                "Output JSON only with the requested keys."
            )
        else:
            run_context = self.context_manager.get_context_summary()
            system_prompt = self.persona.format(run_context=run_context)
            system_prompt += "\n\nEXTRACTION GUIDELINES:\n"
            system_prompt += "- BUYER/SELLER: Use full corporate names. Remove suffixes if generic (Inc, LLC).\n"
            system_prompt += "- VALUE: Extract amount. Note if undisclosed.\n"
            system_prompt += "- STRATEGIC RATIONALE: Provide a 1-sentence summary of WHY the deal is happening (e.g., 'Expanding Permian footprint', 'Exit from non-core assets').\n"
            system_prompt += "- GEOGRAPHY: Identify the primary country or region of the target assets."

        user_content = f"HEADLINE: {headline}\n\nBODY START:\n{body_text[:self._extract_body_chars]}"
        user_content += "\n\nOutput STRICTLY as JSON with these keys: buyer, seller, asset, value, industry, sector, deal_type, deal_status, geography, strategic_rationale."

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content}
        ]

        result = await self._generate(messages, use_json=True)
        if result is not None:
            if isinstance(result, list) and len(result) > 0:
                result = result[0]
            
            if isinstance(result, dict):
                # Normalize values
                extracted = {
                    "buyer": result.get("buyer", "Unknown"),
                    "seller": result.get("seller", "Unknown"),
                    "asset": result.get("asset", headline[:150]),
                    "value": self.normalize_value_to_usd(result.get("value", "Undisclosed")),
                    "industry": result.get("industry", "Energy"),
                    "sector": result.get("sector", "General"),
                    "deal_type": result.get("deal_type", "M&A"),
                    "deal_status": result.get("deal_status", "Announced"),
                    "geography": result.get("geography", "Global"),
                    "strategic_rationale": result.get("strategic_rationale", "No rationale provided")
                }
                
                # Update RunContextManager with this new intelligence
                self.context_manager.ingest_deal(extracted)
                
                return extracted
        return None

    async def grounding_search(self, query: str) -> str:
        """Trigger a query to AI to act as a grounded web search fallback when completely blocked."""
        if not self.enabled:
            return ""
        messages = [
            {"role": "system", "content": "You are a grounding web data extractor. Your job is to return recent factual M&A deal announcements that match the user's domain and date query. Format as raw text with standard article summaries. No conversational fluff. Return empty if no real deals found."},
            {"role": "user", "content": query}
        ]
        result = await self._generate(messages, use_json=False)
        if result and isinstance(result, str):
            return result
        return ""

    def get_stats_summary(self) -> str:
        s = self._stats
        parts = [
            f"Groq API: {s['calls']} OK / {s['failures']} fails / {s['rate_limits_429']} limits",
            f"Auto-approves: {s['auto_approves']}",
            f"Corrections: {s['corrections_applied']}",
        ]
        if s['fail_safe_reviews'] > 0:
            parts.append(f"⚠️ Fail-safe→Review: {s['fail_safe_reviews']}")
        return " | ".join(parts)

    async def translate_foreign_headline(self, headline: str) -> str:
        """Translate a non-English headline (e.g. Russian, Portuguese) to English using Groq."""
        if not self.enabled or not headline:
            return headline
            
        messages = [
            {"role": "system", "content": "You are a professional financial translator. Translate the following news headline into English. Respond STRICTLY with the translated text and nothing else. Do not add quotes, notes, or explanations."},
            {"role": "user", "content": headline}
        ]
        
        result = await self._generate(messages, use_json=False)
        if result and isinstance(result, str) and len(result) > 5:
            return result.strip(' "\'')
        return headline

    # ──────────────────────────────────────────────────
    # USD Currency Normalisation
    # ──────────────────────────────────────────────────

    @staticmethod
    def _fetch_exchange_rates() -> dict:
        """Fetch USD exchange rates from open.er-api.com (free, no API key).
        Results are cached globally for the duration of the current run.
        Cache key is today's date so a next-day run always gets fresh rates.
        """
        global _EXCHANGE_RATES_CACHE
        today = datetime.now().strftime("%Y-%m-%d")
        if _EXCHANGE_RATES_CACHE.get("_date") == today:
            return _EXCHANGE_RATES_CACHE
        try:
            import urllib.request
            import json as _json
            url = "https://open.er-api.com/v6/latest/USD"
            with urllib.request.urlopen(url, timeout=5) as resp:
                data = _json.loads(resp.read())
            rates = data.get("rates", {})
            _EXCHANGE_RATES_CACHE = {"_date": today, **rates}
            logger.info(f"Exchange rates fetched ({len(rates)} currencies)")
        except Exception as e:
            logger.warning(f"Exchange rate fetch failed: {e} — using fallback rates")
            # Fallback hardcoded rates (last known, updated March 2026)
            _EXCHANGE_RATES_CACHE = {
                "_date": today,
                "GBP": 0.79, "EUR": 0.92, "CAD": 1.36, "AUD": 1.55,
                "INR": 83.5, "NOK": 10.5, "DKK": 6.9, "SEK": 10.4,
                "CHF": 0.90, "JPY": 151.0, "CNY": 7.25, "SAR": 3.75,
                "AED": 3.67, "BRL": 4.95, "MXN": 17.2, "SGD": 1.35,
                "HKD": 7.82, "KRW": 1330.0, "ZAR": 18.7, "NZD": 1.63,
            }
        return _EXCHANGE_RATES_CACHE

    def normalize_value_to_usd(self, value_str: str) -> str:
        """Convert a value string to a US$ denominated string.

        Handles patterns like:
          - '£500 million', 'EUR 1.2bn', 'C$800m', 'A$2 billion'
          - 'USD 500 million', '$1.5B'
          - Returns 'Undisclosed' if no numeric value is detectable.

        Always prepends 'US$' to indicate the output currency.
        """
        # Ensure value_str is always a string (AI may return int/float)
        value_str = str(value_str) if value_str is not None else ""
        if not value_str or value_str.strip().lower() in ("", "undisclosed", "unknown", "n/a", "tba"):
            return "Undisclosed"

        v = str(value_str).strip()

        # ── Detect currency code / symbol ──
        CURRENCY_MAP = {
            # Symbol / prefix → ISO code
            "gbp": "GBP", "£": "GBP",
            "eur": "EUR", "€": "EUR",
            "cad": "CAD", "c$": "CAD",
            "aud": "AUD", "a$": "AUD",
            "inr": "INR", "₹": "INR",
            "nok": "NOK", "kr": "NOK",
            "sek": "SEK",
            "dkk": "DKK",
            "chf": "CHF",
            "jpy": "JPY", "¥": "JPY",
            "cny": "CNY",
            "sar": "SAR",
            "aed": "AED",
            "brl": "BRL", "r$": "BRL",
            "mxn": "MXN",
            "sgd": "SGD", "s$": "SGD",
            "hkd": "HKD", "hk$": "HKD",
            "nzd": "NZD", "nz$": "NZD",
            "zar": "ZAR", "r": "ZAR",
            # USD (no conversion needed)
            "usd": "USD", "us$": "USD", "$": "USD", "usd$": "USD",
        }

        detected_iso = "USD"  # Default
        v_lower = v.lower()
        # Check longest matches first to avoid 's$' matching before 'sgd'
        for symbol, iso in sorted(CURRENCY_MAP.items(), key=lambda x: -len(x[0])):
            if v_lower.startswith(symbol) or v_lower.endswith(symbol) or f" {symbol}" in v_lower or f"{symbol} " in v_lower:
                detected_iso = iso
                break

        # ── Extract numeric value ──
        v_clean = re.sub(r'[^\d.,BMKbmk]', ' ', v)
        num_match = re.search(r'[\d]{1,3}(?:[,.]\d{3})*(?:[.,]\d+)?', v_clean)
        if not num_match:
            return "Undisclosed"

        try:
            num = float(num_match.group().replace(",", ""))
        except ValueError:
            return "Undisclosed"

        # ── Detect magnitude suffix ──
        v_lower_full = v.lower()
        if "trillion" in v_lower_full or "tn" in v_lower_full:
            num *= 1_000_000_000_000
        elif "billion" in v_lower_full or "bn" in v_lower_full or v_lower_full.rstrip().endswith("b"):
            num *= 1_000_000_000
        elif "million" in v_lower_full or "mn" in v_lower_full or "mm" in v_lower_full or v_lower_full.rstrip().endswith("m"):
            num *= 1_000_000
        elif "thousand" in v_lower_full or "k" in v_lower_full:
            num *= 1_000

        # ── Convert to USD ──
        if detected_iso != "USD":
            rates = self._fetch_exchange_rates()
            rate = rates.get(detected_iso)
            if rate and rate > 0:
                num = num / rate  # Convert foreign → USD

        # ── Format output ──
        if num >= 1_000_000_000_000:
            return f"US${num / 1_000_000_000_000:.2f} trillion"
        elif num >= 1_000_000_000:
            return f"US${num / 1_000_000_000:.2f} billion"
        elif num >= 1_000_000:
            return f"US${num / 1_000_000:.2f} million"
        elif num >= 1_000:
            return f"US${num / 1_000:.2f} thousand"
        else:
            return f"US${num:.2f}"

    # ──────────────────────────────────────────────────
    # AI-Based Deal Classification (Tier 2 Fallback)
    # ──────────────────────────────────────────────────

    async def classify_deal_sector(self, headline: str, body: str = "",
                                    asset: str = "", industry: str = "",
                                    sector: str = "") -> str:
        """Ask Groq to classify a deal into one of the 5 industry sheets.

        Only called when keyword-based classification in config.py is ambiguous.
        Returns one of: Upstream, Midstream, OFS, R&M, P&U.
        Falls back to P&U if AI fails.
        """
        if not self.enabled:
            return "P&U"

        context = f"Headline: {headline[:300]}\nAsset/Target: {asset}\nIndustry: {industry}\nSector: {sector}\nBody excerpt: {body[:800]}"
        messages = [
            {
                "role": "system",
                "content": (
                    "You are an expert energy-sector M&A deal classifier. Your ONLY job is to classify "
                    "the deal into EXACTLY ONE of 5 categories based on the PRIMARY BUSINESS of the "
                    "asset or company being ACQUIRED (the target, not the buyer).\n\n"
                    "CATEGORIES (pick exactly one):\n"
                    "1. Upstream — Oil & gas exploration and production (E&P). Includes: oil/gas fields, "
                    "wells, reserves, acreage, basins (Permian, Eagle Ford, etc.), working interests, "
                    "mineral rights, shale assets, deepwater blocks, farm-ins/farm-outs, crude production.\n"
                    "2. Midstream — Transportation and processing of hydrocarbons (oil/gas). Includes: pipelines, "
                    "gas processing plants, LNG terminals/export/import, fractionators, NGL infrastructure, "
                    "storage facilities, tank terminals, gathering systems, compressor stations, FSRU/FLNG.\n"
                    "3. OFS — Oilfield services and equipment. Includes: drilling contractors, rig companies, "
                    "hydraulic fracturing, pressure pumping, subsea equipment, well services, wireline, "
                    "seismic surveys, EPC contractors, inspection, tubulars, proppant, chemicals.\n"
                    "4. R&M — Refining, petrochemicals, and fuel marketing. Includes: refineries, crackers, "
                    "ethylene/propylene plants, fuel retail/gas stations, lubricants, asphalt, blending, "
                    "downstream marketing, convenience stores at fuel stations.\n"
                    "5. P&U — Power, utilities, renewables, mining, and clean energy. Includes: solar farms, "
                    "wind farms, hydroelectric, nuclear, coal plants, electricity utilities, grid/transmission, "
                    "battery storage, lithium/cobalt/copper/nickel mining, DLE, hydrogen, ammonia, methanol, CCUS, "
                    "EVs, charging infrastructure, geothermal, biomass, fuel cells, helium.\n\n"
                    "DECISION RULES:\n"
                    "- Focus on the TARGET asset/company, NOT the buyer\n"
                    "- 'Oil and gas producer' or 'E&P company' = Upstream\n"
                    "- A company that OPERATES pipelines = Midstream, even if owned by an E&P\n"
                    "- A DRILLING company = OFS, not Upstream\n"
                    "- If the target does BOTH upstream and midstream, pick the one emphasized more\n"
                    "- When truly uncertain, default to P&U\n\n"
                    "Return ONLY JSON: {\"sheet\": \"<Upstream|Midstream|OFS|R&M|P&U>\", "
                    "\"reason\": \"one-line explanation\"}"
                ),
            },
            {"role": "user", "content": context},
        ]

        result = await self._generate(messages, use_json=True)
        if result and isinstance(result, dict):
            sheet = result.get("sheet", "P&U")
            if sheet in ("Upstream", "Midstream", "OFS", "R&M", "P&U"):
                logger.info(f"AI classified → {sheet}: {headline[:60]} (reason: {result.get('reason', 'N/A')[:60]})")
                return sheet
        return "P&U"
